import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_decks_xlsx():
    wb = openpyxl.Workbook()
    
    # 既存のデフォルトシートを削除し、新規追加
    wb.remove(wb.active)
    
    ws = wb.create_sheet(title="シート0")
    ws.views.sheetView[0].showGridLines = True

    # --- パラメータ定義エリア (1行目〜6行目) ---
    ws["A1"] = "山札枚数 (N)"
    ws["B1"] = 40
    ws["C1"] = 39

    ws["A2"] = "実質ドロー枚数 (n)"
    ws["B2"] = 5
    ws["C2"] = 4

    ws["A3"] = "初期配置契約カード数"
    ws["B3"] = 0
    ws["C3"] = 1

    ws["A4"] = "チヒロ枚数 (T)"
    ws["B4"] = "=SUM(B12)"
    ws["C4"] = "=SUM(C12)"

    ws["A5"] = "サーチ枚数 (S)"
    ws["B5"] = 0
    ws["C5"] = "=SUM(C13:C14)"

    ws["A6"] = "ハズレ枚数 (H)"
    ws["B6"] = "=B1-B3-B4-B5"
    ws["C6"] = "=C1-C4-C5"

    # --- ヘッダー定義 (9行目〜10行目) ---
    ws.merge_cells("A9:A10")
    ws.merge_cells("B9:C9")
    ws["A9"] = "CardNo"
    ws["B9"] = "枚数"
    ws["B10"] = "Deck#0"
    ws["C10"] = "Deck#1"

    ws.merge_cells("E9:E10")
    ws.merge_cells("F9:G9")
    ws["E9"] = "CardNo"
    ws["F9"] = "p初期手札で、手札にある確率 (累積条件付き確率)"
    ws["F10"] = "Deck#0"
    ws["G10"] = "Deck#1"

    # カードデータ定義
    cards = [
        ("Init", 0, 0),
        ("BSC32-025 日下 チヒロ (チヒロ)", 3, 3),
        ("BS44-CP01 世界幼竜グラン・ロロ・ドラゴン (幼ロロ)", 0, 3),
        ("BS72-084 キズナフィールド (キズナF)", 0, 2),
        ("BS67-CX03 プチフェニル (プチフェ)", 0, 0),
        ("SJ15-12 なんて古っ代！ファラオくん (ファラオ)", 0, 3),
        ("BS51-XX03 創界神アレックス＝ロロ (アレロロ)", 0, 1),
        ("BS48-064 神海賊船カリュブデス号 -女神顕現- (女神顕現)", 0, 2),
        ("BS45-055 エジットの天使モニファーエル (モニファ)", 0, 3),
        ("BS74-X01 龍神の覇王ジーク・ヤマト・フリード (ヤマト)", 0, 3),
        ("BS74-X06 氷傑の覇王ロード・ドラゴン・グレイザー (グレイザ)", 0, 3),
        ("BS74-X08 光輝の覇王ルナアーク・カグヤ (カグヤ)", 0, 3),
        ("BS74-X10 轟鉄の覇王サイゴード・ゴレム (サイゴ)", 0, 3),
        ("BS74-086 キズナリターン (キズナR)", 0, 0),
        ("BS55-X02 超覇王ロード・ドラゴン・零 (零)", 0, 0),
        ("X013 黄金皇ロード・ドラゴン・インティ (インティ)", 0, 0),
        ("BS54-009 雷の四天王サカターノ・ベア (サカタ)", 0, 0),
        ("その他", 37, 13)
    ]

    start_row = 11
    # データ書き込み (11行目〜28行目)
    for i, (card_no, q0, q1) in enumerate(cards):
        row = start_row + i
        # 表1: デッキ構成
        ws.cell(row=row, column=1, value=card_no)
        ws.cell(row=row, column=2, value=q0)
        ws.cell(row=row, column=3, value=q1)
        
        # 表2: 確率計算
        ws.cell(row=row, column=5, value=f"=A{row}")
        
        formula_deck0 = (
            f"=(COMBIN(B$1-SUM(B$11:B{row-1}), B$2)-"
            f"COMBIN(B$1-SUM(B$11:B{row}), B$2))/COMBIN(B$1, B$2)"
        )
        formula_deck1 = (
            f"=(COMBIN(C$1-SUM(C$11:C{row-1}), C$2)-"
            f"COMBIN(C$1-SUM(C$11:C{row}), C$2))/COMBIN(C$1, C$2)"
        )
        ws.cell(row=row, column=6, value=formula_deck0)
        ws.cell(row=row, column=7, value=formula_deck1)

    # 29行目: テーブル内合計
    total_row = start_row + len(cards)
    ws.cell(row=total_row, column=1, value="合計")
    ws.cell(row=total_row, column=2, value=f"=SUM(B11:B{total_row-1})")
    ws.cell(row=total_row, column=3, value=f"=SUM(C11:C{total_row-1})")
    
    ws.cell(row=total_row, column=5, value="合計 (累積確率)")
    ws.cell(row=total_row, column=6, value=f"=SUM(F11:F{total_row-1})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G11:G{total_row-1})")

    # --- 確率計算・詳細セクション (30行目〜34行目) ---
    ws.cell(row=total_row+1, column=5, value="1回目: 直接チヒロを引く確率 (実質nドロー)")
    ws.cell(row=total_row+1, column=6, value="=F12")  # チヒロは12行目
    ws.cell(row=total_row+1, column=7, value="=G12")

    ws.cell(row=total_row+2, column=5, value="1回目: サーチ効果経由で引く確率 (実質nドロー)")
    ws.cell(row=total_row+2, column=6, value="=0")
    ws.cell(row=total_row+2, column=7, value="=(G13+G14)*(1-COMBIN(C$1-4-3,3)/COMBIN(C$1-4,3))")  # 幼ロロ(13行目)+キズナF(14行目)

    ws.cell(row=total_row+3, column=5, value="1回目ドロー総合成功確率 (p)")
    ws.cell(row=total_row+3, column=6, value=f"=SUM(F{total_row+1}:F{total_row+2})")
    ws.cell(row=total_row+3, column=7, value=f"=SUM(G{total_row+1}:G{total_row+2})")

    # 精密マリガン式の構築 (Deck#1 用)
    p_fail_k0 = f"(COMBIN(C$5,0)*COMBIN(C$6-3,4)/COMBIN(C$1-3,4))*(C$6-3)/(C$1-3)"
    p_fail_k1 = f"(((COMBIN(C$5,1)*COMBIN(C$6-3,3)/COMBIN(C$1-3,4))*((C$6-2)/(C$1-3))+(COMBIN(C$5,0)*COMBIN(C$6-3,4)/COMBIN(C$1-3,4))*(C$5/(C$1-3)))*(COMBIN(C$1-4-C$4,3)/COMBIN(C$1-4,3)))"
    p_fail_k2 = f"(((COMBIN(C$5,2)*COMBIN(C$6-3,2)/COMBIN(C$1-3,4))*((C$6-1)/(C$1-3))+(COMBIN(C$5,1)*COMBIN(C$6-3,3)/COMBIN(C$1-3,4))*((C$5-1)/(C$1-3)))*(COMBIN(C$1-4-C$4,6)/COMBIN(C$1-4,6)))"
    p_fail_k3 = f"(((COMBIN(C$5,3)*COMBIN(C$6-3,1)/COMBIN(C$1-3,4))*(C$6/(C$1-3))+(COMBIN(C$5,2)*COMBIN(C$6-3,2)/COMBIN(C$1-3,4))*((C$5-2)/(C$1-3)))*(COMBIN(C$1-4-C$4,9)/COMBIN(C$1-4,9)))"
    p_fail_k4 = f"(((COMBIN(C$5,4)*COMBIN(C$6-3,0)/COMBIN(C$1-3,4))*((C$6+1)/(C$1-3))+(COMBIN(C$5,3)*COMBIN(C$6-3,1)/COMBIN(C$1-3,4))*((C$5-3)/(C$1-3)))*(COMBIN(C$1-4-C$4,12)/COMBIN(C$1-4,12)))"
    p_fail_k5 = f"(((COMBIN(C$5,4)*COMBIN(C$6-3,0)/COMBIN(C$1-3,4))*((C$5-4)/(C$1-3)))*(COMBIN(C$1-4-C$4,15)/COMBIN(C$1-4,15)))"
    
    formula_deck1_mulligan = f"=1-({p_fail_k0}+{p_fail_k1}+{p_fail_k2}+{p_fail_k3}+{p_fail_k4}+{p_fail_k5})"

    ws.cell(row=total_row+4, column=5, value="マリガン成功確率 (p_mulligan_total - 実質5枚ドロー)")
    ws.cell(row=total_row+4, column=6, value="=1-(COMBIN(B$1-4-B$4,4)/COMBIN(B$1-4,4))*((B$1-4-B$4)/(B$1-4))")
    ws.cell(row=total_row+4, column=7, value=formula_deck1_mulligan)

    ws.cell(row=total_row+5, column=5, value="★最終成功確率 (マリガン判断を考慮した精密計算)")
    ws.cell(row=total_row+5, column=6, value="=(1-COMBIN(B$1-B$4,4)/COMBIN(B$1,4))+(COMBIN(B$1-B$4,4)/COMBIN(B$1,4))*F33")
    ws.cell(row=total_row+5, column=7, value=(
        "=(((1-COMBIN(C$1-C$4,4)/COMBIN(C$1,4))+"
        "((COMBIN(C$1-C$4,4)-COMBIN(C$1-C$4-C$5,4))/COMBIN(C$1,4))*"
        "(1-COMBIN(C$1-4-C$4,3)/COMBIN(C$1-4,3)))+"
        "(COMBIN(C$6,3)/COMBIN(C$1,3))*G33)"
    ))

    # --- スタイリング定義 ---
    font_family = "Meiryo UI"
    param_font = Font(name=font_family, size=9, italic=True, color="595959")
    header_font = Font(name=font_family, size=10, bold=True)
    body_font = Font(name=font_family, size=10)
    total_font = Font(name=font_family, size=10, bold=True)
    highlight_font = Font(name=font_family, size=11, bold=True, color="C00000")

    param_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    calc_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    double_bottom_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='double', color='000000')
    )

    # パラメータエリアのスタイル
    for r in range(1, 7):
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = param_font
            cell.fill = param_fill
            cell.border = thin_border
            if c == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")

    # ヘッダー (9行目〜10行目)
    for r in range(9, 11):
        for c in range(1, 8):
            if c == 4:
                continue
            cell = ws.cell(row=r, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    # データ行 (11行目〜28行目)
    for r in range(start_row, total_row):
        # 表1: デッキ構成
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.border = thin_border
            if c == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
        
        # 表2: 確率計算
        for c in range(5, 8):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.border = thin_border
            if c == 5:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "0.00%"

    # 合計行 (29行目)
    for c in range(1, 4):
        cell = ws.cell(row=total_row, column=c)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = double_bottom_border
        if c == 1:
            cell.alignment = Alignment(horizontal="left")
        else:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "#,##0"

    for c in range(5, 8):
        cell = ws.cell(row=total_row, column=c)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = double_bottom_border
        if c == 5:
            cell.alignment = Alignment(horizontal="left")
        else:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "0.00%"

    # 詳細計算行 (30行目〜33行目)
    for r in range(total_row + 1, total_row + 5):
        for c in range(5, 8):
            cell = ws.cell(row=r, column=c)
            cell.font = total_font
            cell.fill = calc_fill
            cell.border = thin_border
            if c == 5:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "0.00%"

    # 最終成功確率行 (34行目)
    for c in range(5, 8):
        cell = ws.cell(row=total_row+5, column=c)
        cell.font = highlight_font if c > 5 else total_font
        cell.fill = highlight_fill
        cell.border = double_bottom_border
        if c == 5:
            cell.alignment = Alignment(horizontal="left")
        else:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "0.00%"

    # 列幅調整
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    wb.save("Decks.xlsx")
    print("Decks.xlsx successfully generated in unified Sheet0 layout.")

if __name__ == "__main__":
    create_decks_xlsx()
